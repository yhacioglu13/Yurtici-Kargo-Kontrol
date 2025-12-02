"""YurtiÃ§i Kargo gÃ¶nderilerinde maliyet kontrolÃ¼ ve rapor Ã¼retimi."""
from __future__ import annotations

from pathlib import Path

import pandas as pd
import tkinter as tk
from tkinter import messagebox
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from openpyxl.worksheet.page import PageMargins

# BURASI Ã–NEMLÄ°: config kÃ¶kte, helpers app iÃ§inde
from config import MAIL_EXCEL_ROOT, TARIFF_EXCEL_PATH, RESULT_EXCEL_ROOT
from app.helpers import today_str_en, ensure_folder


def load_base_data() -> pd.DataFrame | None:
    """GÃ¼nlÃ¼k birleÅŸtirilmiÅŸ mail excellerini okur."""
    today = today_str_en()
    base_excel_path = (
        Path(MAIL_EXCEL_ROOT)
        / today
        / f"{today} - BirleÅŸtirilmiÅŸ Mail Excelleri.xlsx"
    )

    if not base_excel_path.exists():
        print(f"âŒ Hata: BirleÅŸtirilmiÅŸ Excel bulunamadÄ±! ({base_excel_path})")
        return None

    try:
        df = pd.read_excel(base_excel_path)
        print(f"ğŸ“„ BirleÅŸtirilmiÅŸ Excel yÃ¼klendi: {base_excel_path}")
        return df
    except Exception as exc:
        print(f"âŒ Hata: BirleÅŸtirilmiÅŸ Excel okunamadÄ±! ({exc})")
        return None


def load_tariff() -> pd.DataFrame | None:
    """Tarife (fiyat) tablosunu okur."""
    tariff_path = Path(TARIFF_EXCEL_PATH)

    if not tariff_path.exists():
        print(f"âŒ Hata: Tarife Excel dosyasÄ± bulunamadÄ±! ({tariff_path})")
        return None

    try:
        df_tarife = pd.read_excel(tariff_path, sheet_name="Tarife")
    except Exception as exc:
        print(f"âŒ Hata: Tarife Excel okunamadÄ±! ({exc})")
        return None

    missing_cols = {"YURTÄ°Ã‡Ä° KARGO", "maliyet"} - set(df_tarife.columns)
    if missing_cols:
        print(f"âŒ Tarife Excel eksik sÃ¼tunlar iÃ§eriyor: {missing_cols}")
        return None

    print(f"ğŸ“„ Tarife Excel yÃ¼klendi: {tariff_path}")
    return df_tarife


def hesapla_maliyet_factory(df_tarife: pd.DataFrame):
    """Desi'ye gÃ¶re maliyet hesaplayan fonksiyonu, tarife tablosuna gÃ¶re Ã¼retir."""

    def get_cost_for_step(step: int) -> float | None:
        matches = df_tarife.loc[df_tarife["YURTÄ°Ã‡Ä° KARGO"] == step, "maliyet"].values
        if matches.size == 0:
            print(f"âš ï¸ Tarife satÄ±rÄ± bulunamadÄ± (desi adÄ±mÄ±={step}).")
            return None
        return matches[0]

    def hesapla_maliyet(desi) -> float | None:
        try:
            desi = float(desi)
        except Exception:
            return None

        if desi < 0:
            return None
        elif desi < 1:
            return get_cost_for_step(1)
        elif desi < 4:
            return get_cost_for_step(2)
        elif desi < 6:
            return get_cost_for_step(3)
        elif desi < 11:
            return get_cost_for_step(4)
        elif desi < 16:
            return get_cost_for_step(5)
        elif desi < 21:
            return get_cost_for_step(6)
        elif desi < 26:
            return get_cost_for_step(7)
        elif desi < 31:
            return get_cost_for_step(8)
        elif desi < 36:
            return get_cost_for_step(9)
        elif desi < 41:
            return get_cost_for_step(10)
        elif desi < 46:
            return get_cost_for_step(11)
        elif desi < 51:
            return get_cost_for_step(12)
        else:
            ana = get_cost_for_step(12)
            ek = get_cost_for_step(13)
            if ana is None or ek is None:
                return None
            return ana + (desi - 50) * ek

    return hesapla_maliyet


def tutar_kontrol_et(row) -> str:
    """ÃœrÃ¼n bedeli ile hesaplanan maliyeti karÅŸÄ±laÅŸtÄ±rÄ±r."""
    try:
        urun_bedeli = float(row["ÃœrÃ¼n Bedeli"])
        maliyet = float(row["maliyet"])
    except Exception:
        return "HesaplanamadÄ±"

    if urun_bedeli == 0:
        return "HesaplanamadÄ±"

    fark_orani = abs(urun_bedeli - maliyet) / urun_bedeli
    return "Tutar uygun" if fark_orani <= 0.05 else "Tutar hatalÄ±!!"


def filter_columns(df: pd.DataFrame) -> pd.DataFrame:
    """Rapor iÃ§in gerekli sÃ¼tunlarÄ± seÃ§er."""
    columns_needed = [
        "Fatura NumarasÄ±",
        "GÃ¶nderici MÃ¼ÅŸteri",
        "GÃ¶nderen MÃ¼ÅŸteri Adresi",
        "AlÄ±cÄ± MÃ¼ÅŸteri",
        "Ã‡Ä±kÄ±ÅŸ Birimi",
        "VarÄ±ÅŸ Birimi",
        "VarÄ±ÅŸ Ä°li",
        "Kargo Tipi",
        "Toplam Kargo Adedi",
        "Desi / Kg",
        "ÃœrÃ¼n Bedeli",
        "Ä°rsaliye MatrahÄ±",
        "Kdv",
        "Ä°rsaliye MatrahÄ±+KDV",
        "Mesafe (Km)",
        "Mesafe AÃ§Ä±klamasÄ±",
        "Teslim Alan",
        "maliyet",
        "SonuÃ§1",
    ]

    available_cols = [col for col in columns_needed if col in df.columns]
    missing = set(columns_needed) - set(available_cols)
    if missing:
        print(f"âš ï¸ Rapor iÃ§in eksik sÃ¼tunlar: {missing}")

    return df[available_cols].copy()


def apply_print_settings(sheet) -> None:
    """Sayfa yazdÄ±rma ayarlarÄ±nÄ± (print setup) dÃ¼zenler.

    - A4, yatay (landscape)
    - GeniÅŸliÄŸi tek sayfaya sÄ±ÄŸdÄ±r
    - Ä°lk satÄ±rÄ± her sayfada tekrar et
    - Kenar boÅŸluklarÄ± makul
    - Yatayda ortalanmÄ±ÅŸ
    """
    # A4 + yatay
    sheet.page_setup.orientation = sheet.ORIENTATION_LANDSCAPE
    sheet.page_setup.paperSize = sheet.PAPERSIZE_A4

    # GeniÅŸliÄŸi 1 sayfaya sÄ±ÄŸdÄ±r, yÃ¼kseklik serbest
    sheet.page_setup.fitToWidth = 1
    sheet.page_setup.fitToHeight = 0
    sheet.sheet_properties.pageSetUpPr.fitToPage = True

    # Kenar boÅŸluklarÄ± (inch cinsinden)
    sheet.page_margins = PageMargins(
        left=0.5,
        right=0.5,
        top=0.75,
        bottom=0.75,
        header=0.3,
        footer=0.3,
    )

    # BaÅŸlÄ±k satÄ±rÄ±nÄ± her sayfada tekrar et (1. satÄ±r)
    sheet.print_title_rows = "1:1"

    # Yatayda sayfayÄ± ortala
    sheet.print_options.horizontalCentered = True


def save_per_invoice(df_filtered: pd.DataFrame) -> None:
    """Her fatura iÃ§in ayrÄ± Excel dosyalarÄ± oluÅŸturur, biÃ§imlendirir ve yazdÄ±rma ayarlarÄ±nÄ± yapar."""
    today = today_str_en()
    save_folder = ensure_folder(Path(RESULT_EXCEL_ROOT) / today)

    fatura_kolon_adi = "Fatura NumarasÄ±"
    if fatura_kolon_adi not in df_filtered.columns:
        print(f"âŒ Hata: '{fatura_kolon_adi}' sÃ¼tunu bulunamadÄ±, fatura bazlÄ± kayÄ±t yapÄ±lamÄ±yor.")
        return

    for fatura_no, grup in df_filtered.groupby(fatura_kolon_adi):
        dosya_adi = f"{fatura_no}.xlsx"
        dosya_yolu = save_folder / dosya_adi
        grup.to_excel(dosya_yolu, index=False)

        workbook = load_workbook(dosya_yolu)
        sheet = workbook.active

        # HÃ¼crelerde metin sarma + hizalama
        for row in sheet.iter_rows():
            for cell in row:
                cell.alignment = Alignment(
                    wrap_text=True,
                    horizontal="left",
                    vertical="top",
                )

        # Ä°lk satÄ±r yÃ¼ksekliÄŸi
        sheet.row_dimensions[1].height = 43.2

        # SÃ¼tun geniÅŸliklerini ayarla
        genislikler = {
            "A": 17,
            "B": 24,
            "C": 30,
            "D": 15,
            "E": 15,
            "F": 15,
            "G": 15,
            "H": 15,
            "I": 10,
            "J": 14,
            "K": 10,
            "L": 18,
            "M": 10,
            "N": 30,
            "O": 20,
            "P": 14,
            "Q": 15,
        }
        for sutun, genislik in genislikler.items():
            if sutun in sheet.column_dimensions:
                sheet.column_dimensions[sutun].width = genislik

        # ğŸ“„ YazdÄ±rma ayarlarÄ±nÄ± uygula (A4, landscape, fit to width, header repeat)
        apply_print_settings(sheet)

        workbook.save(dosya_yolu)

    print(f"âœ… TÃ¼m fatura bazlÄ± dosyalar baÅŸarÄ±yla kaydedildi: {save_folder}")


def run_maliyet_kontrol() -> None:
    """Maliyet kontrolÃ¼ yapar, Ã¶zet tabloyu Ã¼retir ve fatura bazlÄ± excelleri kaydeder."""
    df = load_base_data()
    if df is None:
        return

    df_tarife = load_tariff()
    if df_tarife is None:
        return

    # Tip dÃ¶nÃ¼ÅŸÃ¼mleri
    numeric_cols = ["Desi / Kg", "ÃœrÃ¼n Bedeli"]
    for col in numeric_cols:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors="coerce")
        else:
            print(f"âš ï¸ UyarÄ±: '{col}' kolonu bulunamadÄ±, hesaplamalar eksik olabilir.")

    # Maliyet hesapla
    hesapla_maliyet = hesapla_maliyet_factory(df_tarife)
    df["maliyet"] = df["Desi / Kg"].apply(hesapla_maliyet)

    # Tutar kontrolÃ¼
    df["SonuÃ§1"] = df.apply(tutar_kontrol_et, axis=1)

    # FiltrelenmiÅŸ tablo
    df_filtered = filter_columns(df)

    # Konsolda kÄ±sa bir Ã¶n izleme (var olan sÃ¼tunlara gÃ¶re)
    preview_cols = [
        c
        for c in ["Desi / Kg", "ÃœrÃ¼n Bedeli", "maliyet", "SonuÃ§1"]
        if c in df_filtered.columns
    ]
    if preview_cols:
        print(df_filtered[preview_cols].head())
    else:
        print("âš ï¸ Ã–nizleme iÃ§in uygun sÃ¼tun bulunamadÄ±.")

    # Fatura bazlÄ± dosyalar
    save_per_invoice(df_filtered)
